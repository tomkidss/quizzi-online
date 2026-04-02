# =========================
# App.py
# =========================
import os
import io
import time
import uuid
import sqlite3
import random
import string
import socket
import re
import glob
import logging
from logging.handlers import RotatingFileHandler
from datetime import datetime, date, timezone
from typing import Dict, Any, List, Optional
from functools import wraps

from openpyxl import load_workbook, Workbook
from openpyxl.styles.numbers import is_date_format
from openpyxl.utils.datetime import from_excel

import qrcode

from flask import (
    Flask, render_template, request, redirect, url_for, send_file,
    flash, session, Response, jsonify
)
from flask_socketio import SocketIO, emit, join_room
from werkzeug.security import generate_password_hash, check_password_hash


# =========================
# Config
# =========================
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DB_PATH = os.path.join(BASE_DIR, "quiz.db")
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
EXPORT_DIR = os.path.join(BASE_DIR, "exports")

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(EXPORT_DIR, exist_ok=True)
os.makedirs(os.path.join(BASE_DIR, "logs"), exist_ok=True)

# ===== Logging Config =====
logger = logging.getLogger("QuizActivity")
logger.setLevel(logging.INFO)
# Prevent duplicate logs if handler exists
if not logger.handlers:
    log_formatter = logging.Formatter('[%(asctime)s] - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
    log_file = os.path.join(BASE_DIR, "logs", "quiz_activity.log")
    file_handler = RotatingFileHandler(log_file, maxBytes=5 * 1024 * 1024, backupCount=3, encoding='utf-8')
    file_handler.setFormatter(log_formatter)
    
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(log_formatter)
    
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

app = Flask(__name__)
app.secret_key = os.environ.get("QUIZ_SECRET_KEY", "quizzi-secret-key-default-v9-4")

# eventlet recommended (requirements has eventlet)
socketio = SocketIO(app, cors_allowed_origins="*")

# ===== Admin login (Control security) =====
ADMIN_USERNAME = os.environ.get("QUIZ_ADMIN_USER", "admin")
ADMIN_PASSWORD = os.environ.get("QUIZ_ADMIN_PASS", "Admin@123")  # đổi trước khi chạy thật
ADMIN_PASSWORD_HASH = generate_password_hash(ADMIN_PASSWORD)


# =========================
# In-memory runtime state
# =========================
runtime: Dict[str, Dict[str, Any]] = {}
# runtime[room_code] = {
#   status: "lobby"/"running"/"ended",
#   duration: int (fallback),
#   question_index: int,
#   question_session_id: str,
#   question_start_ms: int,
#   time_up: bool,
#   reveal_on: bool,
#   final_on: bool
# }

def flush_connection_logs(room_code: str):
    events = runtime.get(room_code, {}).get("network_events", [])
    if not events: return
    try:
        log_path = os.path.join(BASE_DIR, "logs", f"quiz_connections_{room_code}.log")
        with open(log_path, "a", encoding="utf-8") as f:
            f.write("\n" + "\n".join(events) + "\n")
        # Clear memory array
        runtime[room_code]["network_events"] = []
    except Exception as e:
        logger.exception(f"Failed to flush connection logs for {room_code}: {e}")


# =========================
# Helpers
# =========================
def now_ms() -> int:
    return int(time.time() * 1000)


def gen_room_code(n: int = 6) -> str:
    alphabet = string.ascii_uppercase + string.digits
    return "".join(random.choice(alphabet) for _ in range(n))


def get_lan_ip() -> str:
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


# =========================
# DB helpers
# =========================
def db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    conn.execute("PRAGMA busy_timeout=5000;")
    return conn


def init_db():
    conn = db()
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS rooms (
      room_code TEXT PRIMARY KEY,
      status TEXT NOT NULL,
      duration INTEGER NOT NULL,
      display_code TEXT,
      player_show_question INTEGER NOT NULL DEFAULT 0,
      player_result_notice INTEGER NOT NULL DEFAULT 0,
      final_display_on INTEGER NOT NULL DEFAULT 0,
      display_show_options INTEGER NOT NULL DEFAULT 0,
      allow_outside_allowlist INTEGER NOT NULL DEFAULT 0,
      quiz_mode TEXT NOT NULL DEFAULT 'host_paced',
      created_at TEXT NOT NULL
    )
    """)

    # migration for older DB
    cur.execute("PRAGMA table_info(rooms)")
    cols = [r[1] for r in cur.fetchall()]
    if "display_code" not in cols:
        cur.execute("ALTER TABLE rooms ADD COLUMN display_code TEXT")
    if "player_show_question" not in cols:
        cur.execute("ALTER TABLE rooms ADD COLUMN player_show_question INTEGER NOT NULL DEFAULT 0")
    if "player_result_notice" not in cols:
        cur.execute("ALTER TABLE rooms ADD COLUMN player_result_notice INTEGER NOT NULL DEFAULT 0")
    if "final_display_on" not in cols:
        cur.execute("ALTER TABLE rooms ADD COLUMN final_display_on INTEGER NOT NULL DEFAULT 0")
    if "display_show_options" not in cols:
        cur.execute("ALTER TABLE rooms ADD COLUMN display_show_options INTEGER NOT NULL DEFAULT 0")
    if "allow_outside_allowlist" not in cols:
        cur.execute("ALTER TABLE rooms ADD COLUMN allow_outside_allowlist INTEGER NOT NULL DEFAULT 0")
    if "quiz_mode" not in cols:
        cur.execute("ALTER TABLE rooms ADD COLUMN quiz_mode TEXT NOT NULL DEFAULT 'host_paced'")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS questions (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      room_code TEXT NOT NULL,
      stt TEXT,
      question_text TEXT NOT NULL,
      image_url TEXT DEFAULT '',
      correct TEXT NOT NULL,
      opt_a TEXT NOT NULL,
      opt_b TEXT NOT NULL,
      opt_c TEXT NOT NULL,
      opt_d TEXT NOT NULL,
      duration_sec INTEGER NOT NULL DEFAULT 20
    )
    """)

    # migration for older DB: questions.duration_sec, questions.image_url
    cur.execute("PRAGMA table_info(questions)")
    qcols = [r[1] for r in cur.fetchall()]
    if "duration_sec" not in qcols:
        cur.execute("ALTER TABLE questions ADD COLUMN duration_sec INTEGER NOT NULL DEFAULT 20")
    if "image_url" not in qcols:
        cur.execute("ALTER TABLE questions ADD COLUMN image_url TEXT DEFAULT ''")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS players (
      id TEXT PRIMARY KEY,
      room_code TEXT NOT NULL,
      name TEXT NOT NULL,
      position TEXT NOT NULL,
      unit TEXT NOT NULL,
      joined_at TEXT NOT NULL,
      employee_code TEXT NOT NULL DEFAULT ''
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS deleted_players (
      id TEXT PRIMARY KEY,
      room_code TEXT NOT NULL,
      name TEXT NOT NULL,
      position TEXT NOT NULL,
      unit TEXT NOT NULL,
      joined_at TEXT NOT NULL,
      employee_code TEXT NOT NULL DEFAULT '',
      deleted_at TEXT NOT NULL
    )
    """)

    # migration for older DB: players.position & players.employee_code
    cur.execute("PRAGMA table_info(players)")
    pcols = [r[1] for r in cur.fetchall()]
    if "position" not in pcols and "employee_code" in pcols:
        cur.execute("ALTER TABLE players RENAME COLUMN employee_code TO position")
        if "barcode" in pcols:
            cur.execute("ALTER TABLE players RENAME COLUMN barcode TO employee_code")
        else:
            cur.execute("ALTER TABLE players ADD COLUMN employee_code TEXT NOT NULL DEFAULT ''")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS question_sessions (
      id TEXT PRIMARY KEY,
      room_code TEXT NOT NULL,
      q_index INTEGER NOT NULL,
      question_id INTEGER NOT NULL,
      start_ms INTEGER NOT NULL,
      duration INTEGER NOT NULL,
      ended_at TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS answers (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      room_code TEXT NOT NULL,
      session_id TEXT NOT NULL,
      player_id TEXT NOT NULL,
      selected_original TEXT NOT NULL,
      is_correct INTEGER NOT NULL,
      elapsed_ms INTEGER NOT NULL,
      submitted_ms INTEGER NOT NULL,
      locked INTEGER NOT NULL
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS option_maps (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      room_code TEXT NOT NULL,
      question_id INTEGER NOT NULL,
      player_id TEXT NOT NULL,
      order_str TEXT NOT NULL
    )
    """)
    cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_option_maps_unique ON option_maps(room_code, question_id, player_id)")

    cur.execute("""
    CREATE TABLE IF NOT EXISTS player_allowlist (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      room_code TEXT NOT NULL,
      name TEXT NOT NULL DEFAULT '',
      position TEXT NOT NULL DEFAULT '',
      unit TEXT NOT NULL DEFAULT '',
      employee_code TEXT NOT NULL DEFAULT ''
    )
    """)

    conn.commit()
    conn.close()


def get_room(room_code: str) -> Optional[sqlite3.Row]:
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM rooms WHERE room_code=?", (room_code,))
    r = cur.fetchone()
    conn.close()
    return r


def set_runtime_from_db(room_code: str):
    room = get_room(room_code)
    if not room:
        return
    runtime.setdefault(room_code, {})
    runtime[room_code].update({
        "status": room["status"],
        "duration": int(room["duration"]),
        "question_index": -1,
        "question_session_id": None,
        "question_start_ms": None,
        "time_up": False,
        "reveal_on": False,
        "final_on": int(room["final_display_on"]) == 1 if ("final_display_on" in room.keys()) else False,
        "display_show_options": int(room["display_show_options"]) == 1 if ("display_show_options" in room.keys()) else False,
    })


def list_players(room_code: str) -> List[Dict[str, Any]]:
    conn = db()
    cur = conn.cursor()
    cur.execute("""
      SELECT id, name, position, unit, joined_at, employee_code
      FROM players WHERE room_code=?
      ORDER BY joined_at ASC
    """, (room_code,))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    
    scoreboard = compute_scoreboard(room_code)
    score_map = {s["player_id"]: s["score"] for s in scoreboard}
    for r in rows:
        r["score"] = score_map.get(r["id"], 0)
        
    return rows


def get_not_joined_list(room_code: str) -> List[Dict[str, Any]]:
    """Return allowlist entries not yet joined (no player record in this room)."""
    conn = db()
    cur = conn.cursor()
    cur.execute("""
        SELECT name, position, unit, employee_code
        FROM player_allowlist
        WHERE room_code=? AND employee_code NOT IN (
            SELECT DISTINCT employee_code FROM players WHERE room_code=?
        )
        ORDER BY id ASC
    """, (room_code, room_code))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


def get_player(player_id: str) -> Optional[sqlite3.Row]:
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM players WHERE id=?", (player_id,))
    row = cur.fetchone()
    conn.close()
    return row


def total_questions(room_code: str) -> int:
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) as c FROM questions WHERE room_code=?", (room_code,))
    row = cur.fetchone()
    conn.close()
    return int(row["c"] or 0)


def get_question_by_index(room_code: str, q_index: int) -> Optional[sqlite3.Row]:
    conn = db()
    cur = conn.cursor()
    cur.execute("""
      SELECT * FROM questions
      WHERE room_code=?
      ORDER BY stt ASC, id ASC
      LIMIT 1 OFFSET ?
    """, (room_code, q_index))
    row = cur.fetchone()
    conn.close()
    return row


def create_question_session(room_code: str, q_index: int, question_id: int, duration: int) -> str:
    sid = uuid.uuid4().hex[:12]
    conn = db()
    cur = conn.cursor()
    cur.execute("""
      INSERT INTO question_sessions(id, room_code, q_index, question_id, start_ms, duration, ended_at)
      VALUES(?,?,?,?,?,?,NULL)
    """, (sid, room_code, q_index, question_id, now_ms(), int(duration)))
    conn.commit()
    conn.close()
    return sid


def end_question_session(session_id: str):
    conn = db()
    cur = conn.cursor()
    cur.execute("UPDATE question_sessions SET ended_at=? WHERE id=?", (datetime.now(timezone.utc).isoformat(), session_id))
    conn.commit()
    conn.close()


def get_or_create_option_order(room_code: str, question_id: int, player_id: str) -> str:
    conn = db()
    cur = conn.cursor()
    cur.execute("""
      SELECT order_str FROM option_maps
      WHERE room_code=? AND question_id=? AND player_id=?
    """, (room_code, question_id, player_id))
    row = cur.fetchone()
    if row:
        conn.close()
        return row["order_str"]

    order = ["A", "B", "C", "D"]
    random.shuffle(order)
    order_str = "".join(order)
    
    try:
        cur.execute("""
          INSERT OR IGNORE INTO option_maps(room_code, question_id, player_id, order_str)
          VALUES(?,?,?,?)
        """, (room_code, question_id, player_id, order_str))
        conn.commit()
    except sqlite3.Error as e:
        logger.error(f"DB Insert option_maps error: {e}")

    # Re-read to ensure we return the correct order even if it was inserted by another thread just now
    cur.execute("""
      SELECT order_str FROM option_maps
      WHERE room_code=? AND question_id=? AND player_id=?
    """, (room_code, question_id, player_id))
    row = cur.fetchone()
    conn.close()
    return row["order_str"] if row else order_str


def has_answer(room_code: str, session_id: str, player_id: str) -> bool:
    conn = db()
    cur = conn.cursor()
    cur.execute("""
      SELECT 1 FROM answers
      WHERE room_code=? AND session_id=? AND player_id=? AND locked=1
      LIMIT 1
    """, (room_code, session_id, player_id))
    ok = cur.fetchone() is not None
    conn.close()
    return ok



def get_player_answer_row(room_code: str, session_id: str, player_id: str) -> Optional[sqlite3.Row]:
    """Lấy đáp án của 1 người chơi cho 1 session (nếu có)."""
    conn = db()
    cur = conn.cursor()
    cur.execute("""
      SELECT * FROM answers
      WHERE room_code=? AND session_id=? AND player_id=? AND locked=1
      ORDER BY submitted_ms ASC
      LIMIT 1
    """, (room_code, session_id, player_id))
    row = cur.fetchone()
    conn.close()
    return row


def compute_option_stats(room_code: str, session_id: str) -> Dict[str, Any]:
    """Thống kê realtime số người chọn A/B/C/D theo đáp án gốc + % trên tổng đã trả lời."""
    base = {"A": 0, "B": 0, "C": 0, "D": 0}
    conn = db()
    cur = conn.cursor()
    cur.execute("""
      SELECT selected_original, COUNT(*) as c
      FROM answers
      WHERE room_code=? AND session_id=? AND locked=1
      GROUP BY selected_original
    """, (room_code, session_id))
    for r in cur.fetchall():
        k = (r["selected_original"] or "").strip().upper()
        if k in base:
            base[k] = int(r["c"] or 0)
    conn.close()

    answered = sum(base.values())
    stats = {}
    for k in ["A", "B", "C", "D"]:
        c = int(base[k] or 0)
        pct = (c * 100.0 / answered) if answered > 0 else 0.0
        stats[k] = {"count": c, "pct": round(pct, 2)}
    stats["answered"] = answered
    return stats


def compute_scoreboard(room_code: str) -> List[Dict[str, Any]]:
    """
    Default scoring (giữ nguyên công thức cũ):
    - Mỗi câu trả lời đúng: +1000
    - Bonus tốc độ: (duration_cua_cau - elapsed_seconds) * 10, tối thiểu 0
    Lưu ý: duration lấy theo từng câu (question_sessions.duration).
    """
    conn = db()
    cur = conn.cursor()

    cur.execute("""
      SELECT a.player_id,
             SUM(CASE WHEN a.is_correct=1 THEN 1000 ELSE 0 END) as base,
             SUM(CASE WHEN a.is_correct=1
                      THEN MAX(0, (qs.duration - (a.elapsed_ms/1000.0)) * 10)
                      ELSE 0 END) as bonus
      FROM answers a
      JOIN question_sessions qs ON qs.id = a.session_id
      WHERE a.room_code=?
      GROUP BY a.player_id
    """, (room_code,))
    rows = cur.fetchall()

    cur.execute("""
      SELECT id, name, position, unit, employee_code
      FROM players WHERE room_code=?
    """, (room_code,))
    pmap = {r["id"]: dict(r) for r in cur.fetchall()}

    scores: List[Dict[str, Any]] = []
    for r in rows:
        pid = r["player_id"]
        base = float(r["base"] or 0)
        bonus = float(r["bonus"] or 0)
        total = int(round(base + bonus))
        info = pmap.get(pid, {"name": "Unknown", "position": "", "unit": "", "employee_code": ""})
        scores.append({
            "player_id": pid,
            "name": info.get("name"),
            "position": info.get("position"),
            "unit": info.get("unit"),
            "employee_code": info.get("employee_code"),
            "score": total
        })

    for pid, info in pmap.items():
        if not any(s["player_id"] == pid for s in scores):
            scores.append({
                "player_id": pid,
                "name": info.get("name"),
                "position": info.get("position"),
                "unit": info.get("unit"),
                "employee_code": info.get("employee_code"),
                "score": 0
            })

    scores.sort(key=lambda x: x["score"], reverse=True)
    conn.close()
    return scores


def compute_scoreboard_extended(room_code: str) -> List[Dict[str, Any]]:
    """
    Scoreboard + thống kê:
    - total_questions
    - correct_count
    - accuracy_pct (2 chữ số)
    - score
    """
    total_q = total_questions(room_code)
    scores = compute_scoreboard(room_code)

    conn = db()
    cur = conn.cursor()
    cur.execute("""
      SELECT player_id,
             SUM(CASE WHEN is_correct=1 THEN 1 ELSE 0 END) as correct_count,
             COUNT(*) as answered_count
      FROM answers
      WHERE room_code=? AND locked=1
      GROUP BY player_id
    """, (room_code,))
    agg = {r["player_id"]: {"correct": int(r["correct_count"] or 0), "answered": int(r["answered_count"] or 0)} for r in cur.fetchall()}
    conn.close()

    out: List[Dict[str, Any]] = []
    for s in scores:
        pid = s["player_id"]
        cc = agg.get(pid, {}).get("correct", 0)
        ans = agg.get(pid, {}).get("answered", 0)
        real_total = max(total_q, ans)
        acc = (cc / real_total * 100.0) if real_total > 0 else 0.0
        out.append({
            **s,
            "total_questions": total_q,
            "correct_count": cc,
            "accuracy_pct": round(acc, 2),
        })
    return out


def compute_player_final_stats(room_code: str, player_id: str) -> Dict[str, Any]:
    total_q = total_questions(room_code)
    scores = compute_scoreboard_extended(room_code)
    score_row = next((x for x in scores if x["player_id"] == player_id), None)

    conn = db()
    cur = conn.cursor()
    cur.execute("""
      SELECT
        SUM(CASE WHEN locked=1 THEN 1 ELSE 0 END) as answered_count,
        SUM(CASE WHEN locked=1 AND is_correct=1 THEN 1 ELSE 0 END) as correct_count
      FROM answers
      WHERE room_code=? AND player_id=?
    """, (room_code, player_id))
    r = cur.fetchone()
    conn.close()

    answered = int(r["answered_count"] or 0) if r else 0
    correct = int(r["correct_count"] or 0) if r else 0
    
    real_total = max(total_q, answered)
    wrong = max(0, answered - correct)
    no_answer = max(0, real_total - answered)
    acc = (correct / real_total * 100.0) if real_total > 0 else 0.0

    return {
        "player_id": player_id,
        "name": (score_row.get("name") if score_row else ""),
        "total_questions": total_q,
        "answered": answered,
        "correct": correct,
        "wrong": wrong,
        "no_answer": no_answer,
        "accuracy_pct": round(acc, 2),
        "score": int(score_row.get("score", 0)) if score_row else 0
    }


def compute_question_stats(room_code: str, session_id: str) -> Dict[str, Any]:
    conn = db()
    cur = conn.cursor()

    cur.execute("""
      SELECT COUNT(*) as total,
             SUM(CASE WHEN locked=1 THEN 1 ELSE 0 END) as answered,
             SUM(CASE WHEN locked=1 AND is_correct=1 THEN 1 ELSE 0 END) as correct,
             SUM(CASE WHEN locked=1 AND is_correct=0 THEN 1 ELSE 0 END) as wrong
      FROM answers
      WHERE room_code=? AND session_id=?
    """, (room_code, session_id))
    row = cur.fetchone()
    answered = int(row["answered"] or 0)
    correct = int(row["correct"] or 0)
    wrong = int(row["wrong"] or 0)

    cur.execute("""
      SELECT player_id, elapsed_ms
      FROM answers
      WHERE room_code=? AND session_id=? AND locked=1 AND is_correct=1
      ORDER BY elapsed_ms ASC
      LIMIT 1
    """, (room_code, session_id))
    fastest = cur.fetchone()
    fastest_info = None
    if fastest:
        pid = fastest["player_id"]
        cur.execute("SELECT name, position, unit FROM players WHERE id=?", (pid,))
        pr = cur.fetchone()
        
        fastest_info = {
            "name": pr["name"] if pr else "Unknown",
            "position": pr["position"] if pr else "",
            "unit": pr["unit"] if pr else "",
            "elapsed_ms": int(fastest["elapsed_ms"] or 0)
        }

    conn.close()

    total_players = len(list_players(room_code))
    no_answer = max(0, total_players - answered)

    return {
        "players_total": total_players,
        "answered": answered,
        "correct": correct,
        "wrong": wrong,
        "no_answer": no_answer,
        "fastest_correct": fastest_info
    }


# =========================
# Excel import/export
# =========================
REQUIRED_COLS = ["STT", "Câu hỏi", "Đáp án", "Phương án A", "Phương án B", "Phương án C", "Phương án D", "Thời gian"]
OPTIONAL_COLS = ["Link ảnh"]


def _convert_drive_url_to_direct(url: str) -> str:
    url = url.strip()
    if "drive.google.com/file/d/" in url:
        # Extract the file ID
        m = re.search(r"/file/d/([a-zA-Z0-9_-]+)", url)
        if m:
            file_id = m.group(1)
            return f"https://drive.google.com/thumbnail?id={file_id}&sz=w1000"
    return url


def _cell_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()



def _format_excel_date(dt: datetime, number_format: str) -> str:
    """
    Map Excel date formats to VN-friendly dd/mm/yy or dd/mm/yyyy output.
    If Excel uses 2-digit year => dd/mm/yy; else dd/mm/yyyy.
    """
    fmt = (number_format or "").lower()
    year_2 = ("yy" in fmt) and ("yyyy" not in fmt)
    y = dt.strftime("%y" if year_2 else "%Y")
    return f"{dt.day:02d}/{dt.month:02d}/{y}"

def excel_cell_to_text(cell) -> str:
    """
    FIX chính:
    - Nếu ô được format kiểu Date nhưng value là chuỗi (text) => TRẢ NGUYÊN chuỗi (không ép).
    - Nếu ô date thật (datetime/date) => xuất dd/mm/yy hoặc dd/mm/yyyy (theo number_format).
    - Nếu ô format date nhưng value là số serial => convert from_excel (theo workbook epoch) rồi format.
    - Nếu float nguyên => 1.0 -> "1"
    - Còn lại: str().strip()
    """
    v = cell.value
    if v is None:
        return ""

    nf = getattr(cell, "number_format", "") or ""
    date_like = is_date_format(nf)

    # 1) Value is already a python date/datetime
    if isinstance(v, datetime):
        return _format_excel_date(v, nf)
    if isinstance(v, date):
        dt = datetime(v.year, v.month, v.day)
        return _format_excel_date(dt, nf)

    # 2) Excel serial number with a date number_format
    if date_like and isinstance(v, (int, float)):
        try:
            wb = cell.parent.parent  # Worksheet -> Workbook
            epoch = getattr(wb, "epoch", None)
            dt = from_excel(v, epoch=epoch) if epoch is not None else from_excel(v)
            if isinstance(dt, datetime):
                return _format_excel_date(dt, nf)
        except Exception:
            pass  # fall through

    # 3) IMPORTANT: date format BUT value is text => keep original text
    if date_like and isinstance(v, str):
        return v.strip()

    # 4) float like 10.0 -> 10
    if isinstance(v, float) and v.is_integer():
        return str(int(v))

    return str(v).strip()


def read_questions_from_excel(path: str) -> List[Dict[str, str]]:
    # data_only=False để lấy đúng dữ liệu gốc (không phụ thuộc công thức)
    wb = load_workbook(path, data_only=False)
    ws = wb.active

    headers: List[str] = []
    for c in range(1, ws.max_column + 1):
        headers.append(_cell_str(ws.cell(row=1, column=c).value))

    hmap = {h: idx + 1 for idx, h in enumerate(headers) if h}

    missing = [c for c in REQUIRED_COLS if c not in hmap]
    if missing:
        raise ValueError("Thiếu cột: " + ", ".join(missing))

    rows: List[Dict[str, str]] = []
    for r in range(2, ws.max_row + 1):
        # ignore fully empty row (dùng excel_cell_to_text để thống nhất)
        if all(excel_cell_to_text(ws.cell(row=r, column=c)) == "" for c in range(1, ws.max_column + 1)):
            continue

        item = {
            "STT": excel_cell_to_text(ws.cell(row=r, column=hmap["STT"])),
            "Câu hỏi": excel_cell_to_text(ws.cell(row=r, column=hmap["Câu hỏi"])),
            "Đáp án": excel_cell_to_text(ws.cell(row=r, column=hmap["Đáp án"])),
            "Phương án A": excel_cell_to_text(ws.cell(row=r, column=hmap["Phương án A"])),
            "Phương án B": excel_cell_to_text(ws.cell(row=r, column=hmap["Phương án B"])),
            "Phương án C": excel_cell_to_text(ws.cell(row=r, column=hmap["Phương án C"])),
            "Phương án D": excel_cell_to_text(ws.cell(row=r, column=hmap["Phương án D"])),
            "Thời gian": excel_cell_to_text(ws.cell(row=r, column=hmap["Thời gian"])),
        }
        if "Link ảnh" in hmap:
            item["Link ảnh"] = excel_cell_to_text(ws.cell(row=r, column=hmap["Link ảnh"]))
        rows.append(item)

    return rows


def validate_questions_rows(rows: List[Dict[str, str]]) -> List[str]:
    errs: List[str] = []
    if not rows:
        return ["File không có dòng dữ liệu hợp lệ."]

    for idx, row in enumerate(rows):
        line = idx + 2
        ans = (row.get("Đáp án") or "").strip().upper()
        if ans not in ["A", "B", "C", "D"]:
            errs.append(f"Dòng {line}: Đáp án không hợp lệ ({ans}) - chỉ A/B/C/D")

        if not (row.get("Câu hỏi") or "").strip():
            errs.append(f"Dòng {line}: Câu hỏi trống")

        for opt in ["Phương án A", "Phương án B", "Phương án C", "Phương án D"]:
            if not (row.get(opt) or "").strip():
                errs.append(f"Dòng {line}: {opt} trống")

        # Thời gian (giây) - cho phép 0 (sẽ dùng fallback), nhưng không được âm / không hợp lệ
        dur_raw = (row.get("Thời gian") or "").strip()
        if dur_raw != "":
            try:
                dur_val = int(float(dur_raw))
                if dur_val < 0:
                    errs.append(f"Dòng {line}: Thời gian phải >= 0")
            except Exception:
                errs.append(f"Dòng {line}: Thời gian không hợp lệ")

    return errs


def import_questions(room_code: str, rows: List[Dict[str, str]]):
    conn = db()
    cur = conn.cursor()

    cur.execute("DELETE FROM questions WHERE room_code=?", (room_code,))
    cur.execute("DELETE FROM option_maps WHERE room_code=?", (room_code,))
    cur.execute("DELETE FROM question_sessions WHERE room_code=?", (room_code,))
    cur.execute("DELETE FROM answers WHERE room_code=?", (room_code,))
    conn.commit()

    for r in rows:
        stt = (r["STT"] or "").strip()
        q = (r["Câu hỏi"] or "").strip()
        correct = (r["Đáp án"] or "").strip().upper()
        a = (r["Phương án A"] or "").strip()
        b = (r["Phương án B"] or "").strip()
        c = (r["Phương án C"] or "").strip()
        d = (r["Phương án D"] or "").strip()
        img_url = (r.get("Link ảnh") or "").strip()
        img_url = _convert_drive_url_to_direct(img_url)

        dur_raw = (r.get("Thời gian") or "").strip()
        try:
            dur_val = int(float(dur_raw)) if dur_raw != "" else 0
        except Exception:
            dur_val = 0

        cur.execute("""
          INSERT INTO questions(room_code, stt, question_text, image_url, correct, opt_a, opt_b, opt_c, opt_d, duration_sec)
          VALUES(?,?,?,?,?,?,?,?,?,?)
        """, (room_code, stt, q, img_url, correct, a, b, c, d, dur_val))

    conn.commit()
    conn.close()


def export_results_excel(room_code: str) -> bytes:
    conn = db()
    cur = conn.cursor()

    cur.execute("SELECT * FROM rooms WHERE room_code=?", (room_code,))
    room = cur.fetchone()

    cur.execute("""
      SELECT id as player_id, room_code, name, position, unit, joined_at, employee_code
      FROM players WHERE room_code=?
      ORDER BY joined_at ASC
    """, (room_code,))
    players = [dict(r) for r in cur.fetchall()]

    cur.execute("""
      SELECT id as player_id, room_code, name, position, unit, joined_at, employee_code, deleted_at
      FROM deleted_players WHERE room_code=?
      ORDER BY deleted_at ASC
    """, (room_code,))
    deleted_players = [dict(r) for r in cur.fetchall()]

    cur.execute("""
      SELECT id as question_id, room_code, stt, question_text, image_url, correct, opt_a, opt_b, opt_c, opt_d, duration_sec
      FROM questions WHERE room_code=?
      ORDER BY stt ASC, id ASC
    """, (room_code,))
    questions = [dict(r) for r in cur.fetchall()]

    cur.execute("""
      SELECT id as session_id, room_code, q_index, question_id, start_ms, duration, ended_at
      FROM question_sessions WHERE room_code=?
      ORDER BY q_index ASC
    """, (room_code,))
    sessions = [dict(r) for r in cur.fetchall()]

    cur.execute("""
      SELECT id, room_code, session_id, player_id, selected_original, is_correct, elapsed_ms, submitted_ms, locked
      FROM answers WHERE room_code=?
      ORDER BY submitted_ms ASC
    """, (room_code,))
    answers = [dict(r) for r in cur.fetchall()]

    scores = compute_scoreboard_extended(room_code)
    # answers sheet: add name
    pmap = {p["player_id"]: p for p in players}
    answers_named = []
    for a in answers:
        pid = a.get("player_id")
        a2 = dict(a)
        a2["name"] = (pmap.get(pid, {}).get("name", ""))
        answers_named.append(a2)
    answers = answers_named

    conn.close()

    # V8: Sheet ma trận chi tiết (mỗi người chơi 1 hàng, mỗi câu 1 cột: đúng=1, sai=0)
    # Map session_id -> q_index
    sess_map = {}
    for s in sessions:
        try:
            sess_map[str(s.get("session_id"))] = int(s.get("q_index", -1))
        except Exception:
            continue

    # Map (player_id, q_index) -> is_correct (1/0)
    ans_map = {}
    for a in answers:
        sid = str(a.get("session_id", ""))
        qix = sess_map.get(sid, None)
        if qix is None or qix < 0:
            continue
        pid = str(a.get("player_id", ""))
        ans_map[(pid, int(qix))] = 1 if int(a.get("is_correct", 0) or 0) == 1 else 0

    T = len(questions)
    detail_rows = []
    for p in players:
        pid = str(p.get("player_id", ""))
        row = {
            "player_id": pid,
            "name": p.get("name", ""),
            "employee_code": p.get("employee_code", ""),
            "position": p.get("position", ""),
            "unit": p.get("unit", "")
        }
        for i in range(T):
            key = f"Q{i+1}"
            v = ans_map.get((pid, i), "")
            row[key] = v
        detail_rows.append(row)

    wb = Workbook()

    def write_sheet(name: str, rows: List[Dict[str, Any]]):
        ws = wb.create_sheet(title=name)
        if not rows:
            ws.append(["(empty)"])
            return
        headers = list(rows[0].keys())
        ws.append(headers)
        for r in rows:
            ws.append([r.get(h, "") for h in headers])

    wb.remove(wb.active)

    meta = [{
        "room_code": room_code,
        "status": room["status"] if room else "",
        "duration_fallback_sec": int(room["duration"]) if room else "",
        "player_show_question": int(room["player_show_question"]) if room and ("player_show_question" in room.keys()) else 0,
        "final_display_on": int(room["final_display_on"]) if room and ("final_display_on" in room.keys()) else 0,
        "exported_at": datetime.now(timezone.utc).isoformat()
    }]
    write_sheet("META", meta)
    write_sheet("PLAYERS", players)
    write_sheet("DELETED_PLAYERS", deleted_players)
    write_sheet("QUESTIONS", questions)
    write_sheet("SESSIONS", sessions)
    write_sheet("ANSWERS", answers)
    write_sheet("DETAIL_MATRIX", detail_rows)
    write_sheet("SCOREBOARD", scores)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def auto_save_excel(room_code: str):
    try:
        room_code = room_code.strip().upper()
        data = export_results_excel(room_code)
        
        save_dir = os.path.join(app.root_path, "exports")
        os.makedirs(save_dir, exist_ok=True)
        
        filename = f"RESULT_{room_code}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
        save_path = os.path.join(save_dir, filename)
        
        with open(save_path, "wb") as f:
            f.write(data)
            
        logger.info(f"Room {room_code} - Auto-saved Excel report to {save_path}")
    except Exception as e:
        logger.exception(f"Room {room_code} - Cannot auto-save Excel: {e}")


# =========================
# Auth decorators
# =========================
def login_required(f):
    @wraps(f)
    def _wrap(*args, **kwargs):
        if not session.get("is_admin"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return _wrap


# =========================
# Routes
# =========================
@app.after_request
def add_header(response):
    response.cache_control.no_cache = True
    response.cache_control.no_store = True
    response.cache_control.must_revalidate = True
    response.headers['Expires'] = '0'
    response.headers['Pragma'] = 'no-cache'
    return response


@app.route("/api/bg/<bg_type>")
def get_dynamic_bg(bg_type):
    if bg_type not in ("display", "player"):
        return jsonify({"url": None})
    pattern = os.path.join(app.static_folder, "backgrounds", f"bg_{bg_type}.*")
    matches = glob.glob(pattern)
    if not matches:
        return jsonify({"url": None})
    # Lấy file đầu tiên tìm thấy
    match = matches[0]
    filename = os.path.basename(match)
    try:
        mtime = int(os.path.getmtime(match))
    except:
        mtime = int(time.time())
    url = url_for("static", filename=f"backgrounds/{filename}", v=mtime)
    return jsonify({"url": url})


@app.route("/")
def home():
    return redirect(url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        if username == ADMIN_USERNAME and check_password_hash(ADMIN_PASSWORD_HASH, password):
            session["is_admin"] = True
            logger.warning("Admin (Control) login successful.")
            return redirect(url_for("control"))
        flash("Sai tài khoản/mật khẩu", "error")
        return redirect(url_for("login"))
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/control")
@login_required
def control():
    ip = get_lan_ip()
    return render_template("control.html", lan_ip=ip)


@app.route("/display")
def display():
    ip = get_lan_ip()
    logo_path = os.path.join(BASE_DIR, "static", "backgrounds", "logo.png")
    logo_exists = os.path.isfile(logo_path)
    return render_template("display.html", lan_ip=ip, logo_exists=logo_exists)


@app.route("/display_gate")
def display_gate():
    ip = get_lan_ip()
    return render_template("display_gate.html", lan_ip=ip)


@app.route("/mc_gate")
def mc_gate():
    ip = get_lan_ip()
    return render_template("mc_gate.html", lan_ip=ip)


@app.route("/mc")
def mc():
    ip = get_lan_ip()
    return render_template("mc.html", lan_ip=ip)

@app.route("/player_join")
def player_join_page():
    return render_template("player_join.html")


@app.route("/player")
def player():
    return render_template("player.html")


@app.route("/room/create", methods=["POST"])
@login_required
def create_room():
    duration = int(request.form.get("duration", "20") or 20)
    duration = max(5, min(duration, 600))
    room_code = gen_room_code(6)
    display_code = uuid.uuid4().hex[:6].upper()

    conn = db()
    cur = conn.cursor()
    cur.execute("""
      INSERT INTO rooms (room_code, status, duration, display_code, player_show_question, final_display_on, display_show_options, created_at)
      VALUES (?,?,?,?,0,0,0,?)
    """, (room_code, "lobby", duration, display_code, datetime.now(timezone.utc).isoformat()))
    conn.commit()
    conn.close()

    logger.warning(f"Admin CREATED NEW ROOM: {room_code} (Duration: {duration}s)")

    set_runtime_from_db(room_code)
    flash(f"Tạo phòng {room_code} thành công.", "success")
    flash(room_code, "new_room_code")
    return redirect(url_for("control"))


@app.route("/room/upload_questions/<room_code>", methods=["POST"])
@login_required
def upload_questions(room_code):
    room_code = room_code.strip().upper()
    if "file" not in request.files:
        flash("Chưa chọn file Excel.", "error")
        return redirect(url_for("control"))

    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xlsm")):
        flash("Vui lòng upload file .xlsx/.xlsm", "error")
        return redirect(url_for("control"))

    save_path = os.path.join(UPLOAD_DIR, f"{room_code}_{int(time.time())}_{f.filename}")
    f.save(save_path)

    try:
        try:
            rows = read_questions_from_excel(save_path)
            errs = validate_questions_rows(rows)
            if errs:
                flash("Lỗi file Excel: " + " | ".join(errs[:10]), "error")
                return redirect(url_for("control"))

            import_questions(room_code, rows)
            logger.info(f"Room {room_code} - HOST uploaded questions (Count: {len(rows)})")

            # reset runtime question index
            set_runtime_from_db(room_code)
            runtime[room_code]["question_index"] = -1
            runtime[room_code]["question_session_id"] = None
            runtime[room_code]["question_start_ms"] = None
            runtime[room_code]["time_up"] = False
            runtime[room_code]["reveal_on"] = False

            # giữ trạng thái hiển thị phương án trên Display theo DB
            _room = get_room(room_code)
            runtime[room_code]["display_show_options"] = (
                (int(_room["display_show_options"]) == 1)
                if (_room and ("display_show_options" in _room.keys()))
                else False
            )

            flash("Upload câu hỏi thành công.", "success")

        except Exception as e:
            logger.exception(f"Room {room_code} - Error reading questions Excel: {e}")
            flash(f"Lỗi đọc Excel: {e}", "error")
    finally:
        if os.path.exists(save_path):
            try:
                os.remove(save_path)
            except Exception as e:
                print(f"Lỗi: Không thể xóa file tạm {save_path}: {e}")

    return redirect(url_for("control"))


@app.route("/room/upload_allowlist/<room_code>", methods=["POST"])
@login_required
def upload_allowlist(room_code):
    room_code = room_code.strip().upper()
    if "file" not in request.files:
        flash("Chưa chọn file Excel.", "error")
        return redirect(url_for("control"))

    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xlsm")):
        flash("Vui lòng upload file .xlsx/.xlsm", "error")
        return redirect(url_for("control"))

    save_path = os.path.join(UPLOAD_DIR, f"allowlist_{room_code}_{int(time.time())}_{f.filename}")
    f.save(save_path)

    try:
        wb = load_workbook(save_path, data_only=True)
        ws = wb.active
        headers = [_cell_str(ws.cell(row=1, column=c).value).lower() for c in range(1, ws.max_column + 1)]

        name_idx = next((i for i, h in enumerate(headers) if h in ("name", "họ tên", "tên")), None)
        position_idx = next((i for i, h in enumerate(headers) if h in ("position", "chức danh")), None)
        unit_idx = next((i for i, h in enumerate(headers) if h in ("unit", "đơn vị")), None)
        ec_idx = next((i for i, h in enumerate(headers) if h in ("employee_code", "mã nv", "mã dự thi")), None)

        if ec_idx is None:
            flash("File thiếu cột employee_code.", "error")
            return redirect(url_for("control"))

        conn = db()
        cur = conn.cursor()
        cur.execute("DELETE FROM player_allowlist WHERE room_code=?", (room_code,))

        count = 0
        for r in range(2, ws.max_row + 1):
            ec = _cell_str(ws.cell(row=r, column=ec_idx + 1).value)
            if not ec:
                continue
            name = _cell_str(ws.cell(row=r, column=name_idx + 1).value) if name_idx is not None else ""
            position = _cell_str(ws.cell(row=r, column=position_idx + 1).value) if position_idx is not None else ""
            unit = _cell_str(ws.cell(row=r, column=unit_idx + 1).value) if unit_idx is not None else ""
            cur.execute(
                "INSERT INTO player_allowlist(room_code, name, position, unit, employee_code) VALUES(?,?,?,?,?)",
                (room_code, name, position, unit, ec)
            )
            count += 1

        conn.commit()
        conn.close()
        logger.info(f"Room {room_code} - HOST uploaded allowlist (Count: {count})")
        flash(f"Upload danh sách kiểm soát thành công: {count} người.", "success")

    except Exception as e:
        logger.exception(f"Room {room_code} - Error reading allowlist: {e}")
        flash(f"Lỗi đọc file allowlist: {e}", "error")
    finally:
        if os.path.exists(save_path):
            try:
                os.remove(save_path)
            except Exception:
                pass

    return redirect(url_for("control"))


@app.route("/room/clear_allowlist/<room_code>", methods=["POST"])
@login_required
def clear_allowlist(room_code):
    room_code = room_code.strip().upper()
    conn = db()
    cur = conn.cursor()
    cur.execute("DELETE FROM player_allowlist WHERE room_code=?", (room_code,))
    conn.commit()
    conn.close()
    logger.warning(f"Room {room_code} - HOST cleared allowlist (Room is now OPEN)")
    flash("Đã xóa danh sách kiểm soát. Phòng trở về chế độ mở.", "success")
    return redirect(url_for("control"))


@app.route("/room/export/<room_code>")
@login_required
def export_room(room_code):
    room_code = room_code.strip().upper()
    data = export_results_excel(room_code)
    filename = f"RESULT_{room_code}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        io.BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


@app.route("/room/reset/<room_code>", methods=["POST"])
@login_required
def reset_room(room_code):
    room_code = room_code.strip().upper()
    conn = db()
    cur = conn.cursor()
    cur.execute("DELETE FROM questions WHERE room_code=?", (room_code,))
    cur.execute("DELETE FROM players WHERE room_code=?", (room_code,))
    cur.execute("DELETE FROM deleted_players WHERE room_code=?", (room_code,))
    cur.execute("DELETE FROM question_sessions WHERE room_code=?", (room_code,))
    cur.execute("DELETE FROM answers WHERE room_code=?", (room_code,))
    cur.execute("DELETE FROM option_maps WHERE room_code=?", (room_code,))
    cur.execute("UPDATE rooms SET status='lobby', final_display_on=0 WHERE room_code=?", (room_code,))
    conn.commit()
    conn.close()

    logger.warning(f"Room {room_code} - HOST RESET room data entirely.")

    runtime.pop(room_code, None)
    flash(f"Đã reset phòng {room_code}.", "success")
    return redirect(url_for("control"))


def refresh_runtime_questions(room_code: str):
    if room_code not in runtime:
        return
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM questions WHERE room_code=? ORDER BY stt ASC, id ASC", (room_code,))
    rows = cur.fetchall()
    conn.close()
    runtime[room_code]["questions"] = [dict(r) for r in rows]



@app.route("/api/questions/<room_code>")
@login_required
def api_get_questions(room_code):
    room_code = room_code.strip().upper()
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT id, stt, question_text, image_url, correct, opt_a, opt_b, opt_c, opt_d, duration_sec FROM questions WHERE room_code=? ORDER BY stt ASC, id ASC", (room_code,))
    rows = cur.fetchall()
    conn.close()
    
    questions = []
    for r in rows:
        questions.append({
            "id": r["id"],
            "stt": r["stt"],
            "question_text": r["question_text"],
            "image_url": r["image_url"],
            "correct": r["correct"],
            "opt_a": r["opt_a"],
            "opt_b": r["opt_b"],
            "opt_c": r["opt_c"],
            "opt_d": r["opt_d"],
            "duration_sec": r["duration_sec"]
        })
    return jsonify({"status": "success", "data": questions})

@app.route("/api/questions/<room_code>/reorder", methods=["POST"])
@login_required
def api_reorder_questions(room_code):
    room_code = room_code.strip().upper()
    data = request.json
    if not data or "ordered_ids" not in data:
        return jsonify({"status": "error", "message": "Missing ordered_ids payload"}), 400
    
    ordered_ids = data["ordered_ids"]
    conn = db()
    cur = conn.cursor()
    try:
        # Bulk update STT from 1 to N mapped over ordered_ids
        for index, q_id in enumerate(ordered_ids):
            cur.execute("UPDATE questions SET stt=? WHERE id=? AND room_code=?", (index + 1, q_id, room_code))
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({"status": "error", "message": f"DB Error: {str(e)}"}), 500
    finally:
        conn.close()
        
    refresh_runtime_questions(room_code)
    return jsonify({"status": "success"})


@app.route("/api/questions/<room_code>/save", methods=["POST"])
@login_required
def api_save_question(room_code):
    room_code = room_code.strip().upper()
    
    q_id = request.form.get("id", "").strip()
    stt = request.form.get("stt", "").strip()
    q_text = request.form.get("question_text", "").strip()
    image_url = request.form.get("image_url", "").strip()
    correct = request.form.get("correct", "A").strip()
    opt_a = request.form.get("opt_a", "").strip()
    opt_b = request.form.get("opt_b", "").strip()
    opt_c = request.form.get("opt_c", "").strip()
    opt_d = request.form.get("opt_d", "").strip()
    try:
        duration_sec = int(request.form.get("duration_sec", 20))
    except ValueError:
        duration_sec = 20
        
    # Handle optional image file upload directly to static/uploads/
    if "image_file" in request.files:
        f = request.files["image_file"]
        if f and f.filename != "":
            upload_dir = os.path.join(BASE_DIR, "static", "uploads")
            os.makedirs(upload_dir, exist_ok=True)
            saved_name = f"{room_code}_{int(time.time())}_{f.filename}"
            save_path = os.path.join(upload_dir, saved_name)
            f.save(save_path)
            # Web accessible URL
            image_url = f"/static/uploads/{saved_name}"

    conn = db()
    cur = conn.cursor()
    try:
        if q_id:
            # Update existing
            cur.execute("""
                UPDATE questions 
                SET stt=?, question_text=?, image_url=?, correct=?, opt_a=?, opt_b=?, opt_c=?, opt_d=?, duration_sec=?
                WHERE id=? AND room_code=?
            """, (stt, q_text, image_url, correct, opt_a, opt_b, opt_c, opt_d, duration_sec, q_id, room_code))
        else:
            # Auto-assign STT to append at the bottom if left blank
            if not stt:
                cur.execute("SELECT MAX(CAST(stt AS INTEGER)) FROM questions WHERE room_code=?", (room_code,))
                max_stt = cur.fetchone()[0]
                stt = str(max_stt + 1) if max_stt is not None else "1"
                
            # Insert new
            cur.execute("""
                INSERT INTO questions (room_code, stt, question_text, image_url, correct, opt_a, opt_b, opt_c, opt_d, duration_sec)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (room_code, stt, q_text, image_url, correct, opt_a, opt_b, opt_c, opt_d, duration_sec))
        conn.commit()
    except Exception as e:
        conn.close()
        return jsonify({"status": "error", "message": str(e)})
        
    conn.close()
    
    # Reload questions array WITHOUT destroying active quiz indices (Fixes Reset & Duplication bug)
    refresh_runtime_questions(room_code)
    
    return jsonify({"status": "success", "message": "Save successful"})


@app.route("/api/questions/<room_code>/delete", methods=["POST"])
@login_required
def api_delete_question(room_code):
    room_code = room_code.strip().upper()
    req = request.get_json()
    q_id = req.get("id") if req else None
    
    if not q_id:
        return jsonify({"status": "error", "message": "Missing Question ID"})
        
    conn = db()
    cur = conn.cursor()
    cur.execute("DELETE FROM questions WHERE id=? AND room_code=?", (q_id, room_code))
    conn.commit()
    conn.close()
    
    # Reload questions array WITHOUT destroying active quiz indices (Fixes Reset & Duplication bug)
    refresh_runtime_questions(room_code)
    
    return jsonify({"status": "success"})


@app.route("/qr/<room_code>.png")
def qr(room_code):
    room_code = room_code.strip().upper()
    ip = get_lan_ip()
    url = f"{request.url_root.replace('http://', 'https://' if 'onrender.com' in request.url_root else 'http://')}player_join?room={room_code}"
    img = qrcode.make(url)
    bio = io.BytesIO()
    img.save(bio, format="PNG")
    bio.seek(0)
    return Response(bio.getvalue(), mimetype="image/png")


@app.route("/mc_qr/<room_code>.png")
def mc_qr(room_code):
    room_code = room_code.strip().upper()
    ip = get_lan_ip()
    url = f"{request.url_root.replace('http://', 'https://' if 'onrender.com' in request.url_root else 'http://')}mc_gate?room={room_code}"
    img = qrcode.make(url)
    bio = io.BytesIO()
    img.save(bio, format="PNG")
    bio.seek(0)
    return Response(bio.getvalue(), mimetype="image/png")


# =========================
# SocketIO: joins
# =========================
@socketio.on("control_join")
def control_join(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    if not room_code:
        return
    join_room(f"room:{room_code}:control")
    emit("control_state", build_control_state(room_code), room=request.sid)


@socketio.on("control_toggle_lock")
def control_toggle_lock(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    if not room_code:
        return
    rt = runtime.setdefault(room_code, {})
    current_lock = rt.get("is_locked", False)
    rt["is_locked"] = not current_lock
    emit_control_and_mc_state(room_code)


@socketio.on("display_join")
def display_join(data):
    room_code = str(data.get("room_code", "")).strip().upper()
    if not room_code:
        return
    join_room(f"room:{room_code}:display")
    emit("display_state", build_display_state(room_code), room=request.sid)
    # Restore QR panel state so refreshing Display doesn't lose the state
    rt = runtime.get(room_code, {})
    if rt.get("display_show_qr"):
        emit("show_qr", {}, room=request.sid)
    else:
        emit("hide_qr", {}, room=request.sid)


@socketio.on("mc_join")
def mc_join(data):
    room_code = str((data or {}).get("room_code", "")).strip().upper()
    code = str((data or {}).get("code", "")).strip()
    if not room_code:
        emit("mc_denied", {"msg": "Thiếu ROOM."}, room=request.sid)
        return

    room = get_room(room_code)
    if not room:
        emit("mc_denied", {"msg": "ROOM không tồn tại."}, room=request.sid)
        return

    expected = str(room["display_code"] or "").strip()
    if (not expected) or (code != expected):
        emit("mc_denied", {"msg": "Sai MC code (dùng chung Display code)."}, room=request.sid)
        return

    join_room(f"room:{room_code}:mc")
    emit("mc_state", build_control_state(room_code), room=request.sid)


@socketio.on("disconnect")
def handle_disconnect():
    logger.debug(f"Client disconnected - SID: {request.sid}")
    sock_info = runtime.get("_sockets", {}).pop(request.sid, None)
    if sock_info:
        r_code = sock_info.get("room_code")
        p_id = sock_info.get("player_id")
        if r_code and p_id:
            now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            runtime.setdefault(r_code, {}).setdefault("network_events", []).append(
                f"[{now_str}] DISCONNECT - Player {p_id} (SID: {request.sid})"
            )

@socketio.on("player_join")
def player_join_evt(data):
    room_code = str(data.get("room_code", "")).strip().upper()
    name = str(data.get("name", "")).strip()
    employee_code = str(data.get("employee_code", "")).strip()
    position = str(data.get("position", "")).strip()
    unit = str(data.get("unit", "")).strip()
    
    # Lấy player_id từ client gửi lên (nếu có)
    client_player_id = str(data.get("player_id", "")).strip()

    if not (room_code and name and employee_code and position and unit):
        emit("join_result", {"ok": False, "msg": "Thiếu thông tin."}, room=request.sid)
        return

    room = get_room(room_code)
    if not room:
        emit("join_result", {"ok": False, "msg": "Phòng không tồn tại."}, room=request.sid)
        return
        
    # Check if room is locked
    rt = runtime.get(room_code, {})
    if rt.get("is_locked", False):
        emit("join_result", {"ok": False, "msg": "Phòng thi đang khóa tham gia, vui lòng liên hệ với Ban tổ chức"}, room=request.sid)
        return
        
    if room["status"] != "lobby" and room["status"] != "running":
        emit("join_result", {"ok": False, "msg": "Phòng đã kết thúc."}, room=request.sid)
        return

    conn = db()
    cur = conn.cursor()

    # Allowlist check: only if the room has an allowlist uploaded
    cur.execute("SELECT COUNT(*) as c FROM player_allowlist WHERE room_code=?", (room_code,))
    allowlist_count = int((cur.fetchone() or {"c": 0})["c"] or 0)
    
    # Bật cờ cho phép ngoài danh sách
    allow_outside = int(dict(room).get("allow_outside_allowlist", 0)) == 1
    
    if allowlist_count > 0 and not allow_outside:
        cur.execute(
            "SELECT 1 FROM player_allowlist WHERE room_code=? AND employee_code=? LIMIT 1",
            (room_code, employee_code)
        )
        if cur.fetchone() is None:
            conn.close()
            emit("join_result", {"ok": False, "msg": "Không có Mã NV/Mã dự thi này trong Danh sách người chơi được tham gia"}, room=request.sid)
            return

    # Check for existing player by employee_code in this room
    cur.execute("SELECT id FROM players WHERE room_code=? AND employee_code=?", (room_code, employee_code))
    existing = cur.fetchone()

    player_id = ""

    if existing:
        db_player_id = existing[0]
        # Nếu Client gửi lên một ID trống, tức là họ muốn tạo mới, NHƯNG mã số này đã có người dùng
        if not client_player_id:
            conn.close()
            emit("join_result", {"ok": False, "msg": f"Mã NV/Mã dự thi '{employee_code}' đã được đăng ký bởi người khác."}, room=request.sid)
            return
        
        # Nếu Client có gửi ID, mà ID đó không khớp ID trong DB thì cũng chặn
        if client_player_id != db_player_id:
            conn.close()
            emit("join_result", {"ok": False, "msg": f"Mã NV/Mã dự thi '{employee_code}' đã được đăng ký bởi người khác."}, room=request.sid)
            return

        # Tới đây thì client_player_id khớp với db_player_id -> Cho phép Reconnect
        player_id = db_player_id
        
        # Update existing player info just in case they changed name/unit
        cur.execute("""
            UPDATE players 
            SET name=?, position=?, unit=?
            WHERE id=?
        """, (name, position, unit, player_id))
    else:
        player_id = uuid.uuid4().hex[:10]
        cur.execute("""
          INSERT INTO players(id, room_code, name, position, unit, joined_at, employee_code)
          VALUES(?,?,?,?,?,?,?)
        """, (player_id, room_code, name, position, unit, datetime.now(timezone.utc).isoformat(), employee_code))
        
    conn.commit()
    conn.close()

    logger.info(f"Room {room_code} - Player Joined: {name} (EmpCode: {employee_code}, Pos: {position}, Unit: {unit}) -> ID: {player_id}")

    # Deferred connection logging
    runtime.setdefault("_sockets", {})[request.sid] = {"room_code": room_code, "player_id": player_id}
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    runtime.setdefault(room_code, {}).setdefault("network_events", []).append(
        f"[{now_str}] JOIN - Player {player_id} (SID: {request.sid})"
    )

    join_room(f"room:{room_code}:players")
    join_room(f"player:{player_id}")
    emit("join_result", {"ok": True, "player_id": player_id, "room_code": room_code})

    socketio.emit("player_list_updated", {"players": list_players(room_code)}, room=f"room:{room_code}:control")
    socketio.emit("player_list_updated", {"players": list_players(room_code)}, room=f"room:{room_code}:display")
    socketio.emit("player_list_updated", {"players": list_players(room_code)}, room=f"room:{room_code}:mc")

    # Emit updated "not joined" list if there's an allowlist
    not_joined = get_not_joined_list(room_code)
    socketio.emit("not_joined_updated", {"players": not_joined}, room=f"room:{room_code}:control")
    socketio.emit("not_joined_updated", {"players": not_joined}, room=f"room:{room_code}:display")


@socketio.on("player_enter_room")
def player_enter_room(data):
    room_code = str(data.get("room_code", "")).strip().upper()
    player_id = str(data.get("player_id", "")).strip()
    if not room_code or not player_id:
        return

    # Nếu player đã bị xóa khỏi DB (hoặc không tồn tại), không cho tiếp tục xem/thi
    if not get_player(player_id):
        emit("player_kicked", {"room_code": room_code, "player_id": player_id, "msg": "Bạn đã bị xóa tên, hãy tham gia lại."}, room=request.sid)
        return

    logger.debug(f"Room {room_code} - Player Re-entered: {player_id}")

    # Deferred connection logging
    runtime.setdefault("_sockets", {})[request.sid] = {"room_code": room_code, "player_id": player_id}
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    runtime.setdefault(room_code, {}).setdefault("network_events", []).append(
        f"[{now_str}] RE-ENTER - Player {player_id} (SID: {request.sid})"
    )

    join_room(f"room:{room_code}:players")
    join_room(f"player:{player_id}")
    emit("player_state", build_player_state(room_code, player_id), room=request.sid)


# =========================
# Build states
# =========================
def emit_control_and_mc_state(room_code: str):
    """Emit latest state to both Control and MC rooms."""
    state = build_control_state(room_code)
    socketio.emit("control_state", state, room=f"room:{room_code}:control")
    socketio.emit("mc_state", state, room=f"room:{room_code}:mc")


def build_control_state(room_code: str) -> Dict[str, Any]:
    room = get_room(room_code)
    rt = runtime.get(room_code, {})
    session_id = rt.get("question_session_id")
    q_index = rt.get("question_index", -1)

    current_question = None
    if q_index is not None and q_index >= 0:
        q = get_question_by_index(room_code, q_index)
        if q:
            current_question = {
                "id": int(q["id"]),
                "stt": q["stt"],
                "question_text": q["question_text"],
                "correct": q["correct"],
                "opt_a": q["opt_a"],
                "opt_b": q["opt_b"],
                "opt_c": q["opt_c"],
                "opt_d": q["opt_d"],
                "image_url": q["image_url"] if "image_url" in q.keys() else "",
                "duration_sec": int(q["duration_sec"]) if ("duration_sec" in q.keys()) else None,
                "start_ms": rt.get("question_start_ms")
            }

    stats = compute_question_stats(room_code, session_id) if session_id else None
    option_stats = compute_option_stats(room_code, session_id) if session_id else None
    top10 = compute_scoreboard(room_code)[:10]

    room_dict = dict(room) if room else None
    if room_dict and "display_code" not in room_dict:
        # SQLite Row fallback if it wasn't selected or something
        room_dict["display_code"] = room["display_code"] if room else ""

    final_board = []
    if room_dict and room_dict.get("status") == "ended":
        final_board = compute_scoreboard_extended(room_code)
    
    return {
        "room": room_dict,
        "runtime": rt,
        "is_locked": bool(rt.get("is_locked", False)),
        "display_show_options": bool(rt.get("display_show_options", False)),
        "flags": {
            "time_up": bool(rt.get("time_up")),
            "reveal_on": bool(rt.get("reveal_on")),
        },
        "players": list_players(room_code),
        "question_count": total_questions(room_code),
        "current_question": current_question,
        "stats": stats,
        "option_stats": option_stats,
        "top10": top10,
        "final_board": final_board,
        "mc_show_correct_answer": bool(rt.get("mc_show_correct_answer", True))
    }


def build_display_state(room_code: str) -> Dict[str, Any]:
    rt = runtime.get(room_code, {})
    q_index = rt.get("question_index", -1)
    session_id = rt.get("question_session_id")
    
    stats = compute_question_stats(room_code, session_id) if session_id else None

    base_state: Dict[str, Any] = {
        "display_show_player_list": bool(rt.get("display_show_player_list", False)),
        "display_show_winner": bool(rt.get("display_show_winner", False)),
        "players": list_players(room_code) if rt.get("display_show_player_list", False) else [],
        "quiz_mode": rt.get("quiz_mode", "host_paced"),
        "stats": stats
    }
    
    if base_state["quiz_mode"] == "self_paced":
        if rt.get("status") != "running":
            base_state["status"] = rt.get("status", "lobby")
            return base_state
            
        base_state["status"] = "running"
        sessions = rt.get("sessions", [])
        total_q = len(sessions)
        
        scores = compute_scoreboard(room_code)
        
        current_rank = 1
        for i, s in enumerate(scores):
            if i > 0 and s["score"] < scores[i-1]["score"]:
                current_rank = i + 1
            s["rank"] = current_rank
            
            pstate = rt.get("player_states", {}).get(s["player_id"], {})
            s["q_index"] = pstate.get("q_index", 0)
            s["is_finished"] = pstate.get("is_finished", False)
            s["total_questions"] = total_q
            
        base_state["live_leaderboard"] = scores
        return base_state

    if rt.get("status") != "running" or q_index < 0:
        base_state["status"] = rt.get("status", "lobby")
        return base_state

    q = get_question_by_index(room_code, q_index)
    if not q:
        base_state["status"] = rt.get("status", "running")
        return base_state

    base_state["status"] = "running"
    base_state["q_index"] = q_index
    base_state["question_text"] = q["question_text"]
    base_state["image_url"] = q["image_url"] if "image_url" in q.keys() else ""
    base_state["stt"] = q["stt"]
    base_state["start_ms"] = rt.get("question_start_ms")
    base_state["duration"] = rt.get("duration", 20)
    base_state["display_show_options"] = bool(rt.get("display_show_options", False))
    base_state["options_original"] = [q["opt_a"], q["opt_b"], q["opt_c"], q["opt_d"]]
    
    return base_state


def build_player_state(room_code: str, player_id: str) -> Dict[str, Any]:
    rt = runtime.get(room_code, {})
    quiz_mode = rt.get("quiz_mode", "host_paced")
    
    if quiz_mode == "self_paced":
        pstate = rt.get("player_states", {}).get(player_id)
        if not pstate:
            return {"status": "lobby"}
        if pstate.get("is_finished"):
            return {"status": "ended"}
        q_index = pstate.get("q_index", 0)
        sessions = rt.get("sessions", [])
        if q_index >= len(sessions):
            return {"status": "ended"}
    else:
        q_index = rt.get("question_index", -1)
        
    if rt.get("status") != "running" or q_index < 0:
        return {"status": rt.get("status", "lobby")}

    q = get_question_by_index(room_code, q_index)
    if not q:
        return {"status": rt.get("status", "running")}

    qid = int(q["id"])
    order_str = get_or_create_option_order(room_code, qid, player_id)
    order = list(order_str)

    options_map = {"A": q["opt_a"], "B": q["opt_b"], "C": q["opt_c"], "D": q["opt_d"]}
    displayed = [options_map[k] for k in order]

    locked = False
    answered_current = False
    selected_index_saved = None
    last_is_correct = None
    
    if quiz_mode == "self_paced":
        sess = rt.get("sessions", [])[q_index]
        session_id = sess["session_id"]
        start_ms = pstate.get("start_ms", now_ms())
        duration = sess["duration"]
        time_up_flag = False
    else:
        session_id = rt.get("question_session_id")
        start_ms = rt.get("question_start_ms")
        duration = rt.get("duration", 20)
        time_up_flag = bool(rt.get("time_up"))
    deadline_ms = int(start_ms) + int(duration) * 1000 if start_ms else None
    deadline_passed = (deadline_ms is not None and now_ms() > int(deadline_ms))

    if session_id:
        answered_current = has_answer(room_code, session_id, player_id)
        if answered_current:
            row_ans = get_player_answer_row(room_code, session_id, player_id)
            if row_ans:
                sel_orig = (row_ans["selected_original"] or "").strip().upper()
                # also capture correctness for end-of-time notice
                try:
                    last_is_correct = int(row_ans["is_correct"]) if ("is_correct" in row_ans.keys()) else None
                except Exception:
                    last_is_correct = None

                if sel_orig in order:
                    selected_index_saved = order.index(sel_orig)

    locked = bool(answered_current) or bool(time_up_flag) or bool(deadline_passed)

    player = get_player(player_id)
    room = get_room(room_code)
    show_q = int(room["player_show_question"]) == 1 if room and ("player_show_question" in room.keys()) else False

    payload = {
        "status": "running",
        "q_index": q_index,
        "q_number": int(q_index) + 1,
        "player_name": (player["name"] if player else ""),
        "stt": q["stt"],
        "start_ms": start_ms,
        "duration": duration,
        "total_questions": total_questions(room_code),
        "answered_current": bool(answered_current),
        "selected_index": selected_index_saved,
        "last_is_correct": last_is_correct,
        "options": displayed,
        "locked": locked,
        "time_up": bool(time_up_flag) or bool(deadline_passed),
        "result_notice": False if quiz_mode == "self_paced" else (int(room["player_result_notice"])==1 if room and ("player_result_notice" in room.keys()) else False),
        "show_question": show_q
    }
    if show_q:
        payload["question_text"] = q["question_text"]
        payload["image_url"] = q["image_url"] if "image_url" in q.keys() else ""
    return payload


# =========================
# Timer: server time_up
# =========================
def schedule_time_up(room_code: str, session_id: str, duration: int):
    def _task():
        socketio.sleep(duration)
        rt = runtime.get(room_code, {})
        if rt.get("question_session_id") != session_id:
            return
        rt["time_up"] = True
        socketio.emit("time_up", {}, room=f"room:{room_code}:display")
        socketio.emit("time_up", {}, room=f"room:{room_code}:players")
        socketio.emit("time_up", {}, room=f"room:{room_code}:mc")
        socketio.emit("time_up", {}, room=f"room:{room_code}:control")
        emit_control_and_mc_state(room_code)
    socketio.start_background_task(_task)

def self_paced_timer_daemon(room_code: str):
    def _task():
        while True:
            socketio.sleep(1)
            rt = runtime.get(room_code)
            if not rt or rt.get("status") != "running" or rt.get("quiz_mode") != "self_paced":
                break
            
            now = now_ms()
            sessions = rt.get("sessions", [])
            has_changes = False
            
            for pid, pstate in list(rt.get("player_states", {}).items()):
                if pstate.get("is_finished"):
                    continue
                q_index = pstate.get("q_index", 0)
                if q_index >= len(sessions):
                    continue
                
                sess = sessions[q_index]
                duration = sess["duration"]
                start_ms = pstate["start_ms"]
                deadline_ms = start_ms + duration * 1000
                
                if now > deadline_ms:
                    # Time up -> auto advance
                    pstate["q_index"] += 1
                    pstate["start_ms"] = now
                    has_changes = True
                    
                    if pstate["q_index"] >= len(sessions):
                        pstate["is_finished"] = True
                        socketio.emit("quiz_ended", {"room_code": room_code}, room=f"player:{pid}")
                    else:
                        socketio.emit("question_started", {"q_index": pstate["q_index"]}, room=f"player:{pid}")
            
            if has_changes:
                emit_control_and_mc_state(room_code)
    
    socketio.start_background_task(_task)

# =========================
# SocketIO: control actions
# =========================
@socketio.on("player_request_final_stats")
def player_request_final_stats(data):
    # Fallback: when player misses the end-of-quiz push, allow requesting stats explicitly
    room_code = str(data.get("room_code", "")).strip().upper()
    player_id = str(data.get("player_id", "")).strip()
    if not room_code or not player_id:
        return
    try:
        emit("player_final_stats", compute_player_final_stats(room_code, player_id), room=request.sid)
    except Exception as e:
        logger.exception(f"Room {room_code} - Error generating final stats for {player_id}: {e}")
        # silent fail to avoid breaking socket flow
        return


@socketio.on("control_start_quiz")
def control_start_quiz(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    if not room_code:
        return

    if total_questions(room_code) <= 0:
        emit("toast", {"type": "error", "msg": "Chưa có câu hỏi. Vui lòng upload Excel trước."}, room=request.sid)
        return

    logger.info(f"Room {room_code} - HOST started the quiz.")
    conn = db()
    cur = conn.cursor()
    
    cur.execute("SELECT status, quiz_mode, duration FROM rooms WHERE room_code=?", (room_code,))
    row = cur.fetchone()
    if row and row["status"] in ("lobby", "ended"):
        if row["status"] == "ended":
            auto_save_excel(room_code)
        cur.execute("DELETE FROM question_sessions WHERE room_code=?", (room_code,))
        cur.execute("DELETE FROM answers WHERE room_code=?", (room_code,))
        cur.execute("DELETE FROM option_maps WHERE room_code=?", (room_code,))

    cur.execute("UPDATE rooms SET status='running', final_display_on=0 WHERE room_code=?", (room_code,))
    conn.commit()
    conn.close()

    logger.info(f"Room {room_code} - HOST STARTED THE QUIZ.")

    quiz_mode = row["quiz_mode"] if row else "host_paced"
    
    runtime.setdefault(room_code, {})
    runtime[room_code]["status"] = "running"
    runtime[room_code]["quiz_mode"] = quiz_mode
    runtime[room_code]["duration"] = int(row["duration"]) if row else 20
    runtime[room_code]["final_on"] = False
    runtime[room_code]["question_index"] = -1
    runtime[room_code]["question_session_id"] = None
    runtime[room_code]["question_start_ms"] = None
    runtime[room_code]["time_up"] = False
    runtime[room_code]["reveal_on"] = False

    if quiz_mode == "self_paced":
        start_self_paced_quiz(room_code)
    else:
        start_next_question(room_code)


@socketio.on("control_next_question")
def control_next_question(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    logger.info(f"Room {room_code} - HOST clicked Next Question.")
    start_next_question(room_code)


@socketio.on("control_end_quiz")
def control_end_quiz(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()

    # end session if running
    rt = runtime.get(room_code)
    if rt and rt.get("question_session_id"):
        end_question_session(rt["question_session_id"])

    logger.info(f"Room {room_code} - HOST ended the quiz.")
    conn = db()
    cur = conn.cursor()
    cur.execute("UPDATE rooms SET status='ended' WHERE room_code=?", (room_code,))
    conn.commit()
    conn.close()

    # Auto-save Excel on quiz end
    auto_save_excel(room_code)

    # Flush deferred connection logs to disk
    flush_connection_logs(room_code)

    runtime.setdefault(room_code, {})
    runtime[room_code]["status"] = "ended"
    runtime[room_code]["time_up"] = True
    runtime[room_code]["reveal_on"] = False

    socketio.emit("hide_answer", {}, room=f"room:{room_code}:display")
    socketio.emit("hide_stats", {}, room=f"room:{room_code}:display")

    socketio.emit("quiz_ended", {"room_code": room_code}, room=f"room:{room_code}:display")
    socketio.emit("quiz_ended", {"room_code": room_code}, room=f"room:{room_code}:players")
    socketio.emit("quiz_ended", {"room_code": room_code}, room=f"room:{room_code}:mc")
    # gửi thống kê cá nhân cho từng player
    for p in list_players(room_code):
        pid = p.get("id")
        if pid:
            socketio.emit("player_final_stats", compute_player_final_stats(room_code, pid), room=f"player:{pid}")

    emit_control_and_mc_state(room_code)


@socketio.on("control_reveal_answer")
def control_reveal_answer(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    rt = runtime.get(room_code, {})
    if rt.get("status") != "running":
        return
    rt["reveal_on"] = True
    logger.info(f"Room {room_code} - HOST Revealed Answer.")

    q_index = rt.get("question_index", -1)
    q = get_question_by_index(room_code, q_index)
    if not q:
        return
    correct = q["correct"]
    opt_text = {"A": q["opt_a"], "B": q["opt_b"], "C": q["opt_c"], "D": q["opt_d"]}.get(correct, "")

    socketio.emit("reveal_answer", {"text": opt_text}, room=f"room:{room_code}:display")
    socketio.emit("reveal_answer", {"text": opt_text}, room=f"room:{room_code}:players")
    emit_control_and_mc_state(room_code)


@socketio.on("control_hide_answer")
def control_hide_answer(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    rt = runtime.get(room_code, {})
    rt["reveal_on"] = False
    socketio.emit("hide_answer", {}, room=f"room:{room_code}:display")
    socketio.emit("hide_answer", {}, room=f"room:{room_code}:players")
    emit_control_and_mc_state(room_code)


@socketio.on("control_show_stats_on_display")
def control_show_stats_on_display(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    rt = runtime.get(room_code, {})
    sid = rt.get("question_session_id")
    if not sid:
        return
    
    logger.info(f"Room {room_code} - HOST Showed Stats on Display.")
    
    stats = compute_question_stats(room_code, sid)
    top10 = compute_scoreboard(room_code)[:10]
    socketio.emit("show_stats", {"stats": stats, "top10": top10}, room=f"room:{room_code}:display")
    
    full_scoreboard = compute_scoreboard(room_code)
    socketio.emit("show_stats", {"stats": stats, "scoreboard": full_scoreboard}, room=f"room:{room_code}:players")


@socketio.on("control_hide_stats_on_display")
def control_hide_stats_on_display(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    socketio.emit("hide_stats", {}, room=f"room:{room_code}:display")
    socketio.emit("hide_stats", {}, room=f"room:{room_code}:players")


@socketio.on("control_set_player_show_question")
def control_set_player_show_question(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    enabled = bool(data.get("enabled", False))

    room = get_room(room_code)
    if not room:
        return
    # cho phép bật/tắt realtime cả khi đang thi
    conn = db()
    cur = conn.cursor()
    cur.execute("UPDATE rooms SET player_show_question=? WHERE room_code=?", (1 if enabled else 0, room_code))
    conn.commit()
    conn.close()

    logger.info(f"Room {room_code} - HOST toggled Player Show Question: {enabled}")

    emit_control_and_mc_state(room_code)
    socketio.emit("player_config", {"show_question": enabled}, room=f"room:{room_code}:players")



@socketio.on("control_set_quiz_mode")
def control_set_quiz_mode(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    mode = str(data.get("mode", "host_paced")).strip()

    room = get_room(room_code)
    if not room:
        return
    if room["status"] != "lobby":
        # Cannot change mode after starting
        return

    conn = db()
    cur = conn.cursor()
    cur.execute("UPDATE rooms SET quiz_mode=? WHERE room_code=?", (mode, room_code))
    conn.commit()
    conn.close()

    logger.info(f"Room {room_code} - HOST set Quiz Mode: {mode}")
    emit_control_and_mc_state(room_code)

@socketio.on("control_set_allow_outside_allowlist")
def control_set_allow_outside_allowlist(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    enabled = bool(data.get("enabled", False))

    room = get_room(room_code)
    if not room:
        return
    
    conn = db()
    cur = conn.cursor()
    cur.execute("UPDATE rooms SET allow_outside_allowlist=? WHERE room_code=?", (1 if enabled else 0, room_code))
    conn.commit()
    conn.close()

    logger.info(f"Room {room_code} - HOST toggled Allow Outside Allowlist: {enabled}")
    emit_control_and_mc_state(room_code)

@socketio.on("control_set_player_result_notice")
def control_set_player_result_notice(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    enabled = bool(data.get("enabled", False))

    room = get_room(room_code)
    if not room:
        return
    # cho phép bật/tắt realtime cả khi đang thi
    conn = db()
    cur = conn.cursor()
    cur.execute("UPDATE rooms SET player_result_notice=? WHERE room_code=?", (1 if enabled else 0, room_code))
    conn.commit()
    conn.close()

    logger.info(f"Room {room_code} - HOST toggled Player Result Notice: {enabled}")

    emit_control_and_mc_state(room_code)
    socketio.emit("player_config", {"result_notice": enabled}, room=f"room:{room_code}:players")

@socketio.on("control_toggle_final_results_on_display")
def control_toggle_final_results_on_display(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    enabled = bool(data.get("enabled", False))

    room = get_room(room_code)
    if not room:
        return
    if room["status"] != "ended":
        emit("toast", {"type": "warn", "msg": "Chỉ có thể Hiện/Ẩn bảng chung cuộc sau khi bài thi kết thúc."}, room=request.sid)
        return

    conn = db()
    cur = conn.cursor()
    cur.execute("UPDATE rooms SET final_display_on=? WHERE room_code=?", (1 if enabled else 0, room_code))
    conn.commit()
    conn.close()

    logger.info(f"Room {room_code} - HOST toggled Final Results on Display: {enabled}")

    if enabled:
        final_board = compute_scoreboard_extended(room_code)
        socketio.emit("show_final_results", {"rows": final_board}, room=f"room:{room_code}:display")
    else:
        socketio.emit("hide_final_results", {}, room=f"room:{room_code}:display")

    emit_control_and_mc_state(room_code)


@socketio.on("control_show_qr_on_display")
def control_show_qr_on_display(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    runtime.setdefault(room_code, {})["display_show_qr"] = True
    socketio.emit("show_qr", {}, room=f"room:{room_code}:display")


@socketio.on("control_hide_qr_on_display")
def control_hide_qr_on_display(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    runtime.setdefault(room_code, {})["display_show_qr"] = False
    socketio.emit("hide_qr", {}, room=f"room:{room_code}:display")

@socketio.on("control_show_winner_on_display")
def control_show_winner_on_display(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    rt = runtime.setdefault(room_code, {})
    rt["display_show_winner"] = True
    
    sid = rt.get("question_session_id")
    stats = compute_question_stats(room_code, sid) if sid else None
    socketio.emit("show_winner", {"stats": stats}, room=f"room:{room_code}:display")
    socketio.emit("show_winner", {"stats": stats}, room=f"room:{room_code}:players")

@socketio.on("control_hide_winner_on_display")
def control_hide_winner_on_display(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    rt = runtime.setdefault(room_code, {})
    rt["display_show_winner"] = False
    socketio.emit("hide_winner", {}, room=f"room:{room_code}:display")
    socketio.emit("hide_winner", {}, room=f"room:{room_code}:players")

@socketio.on("control_mc_show_answer")
def control_mc_show_answer(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    rt = runtime.setdefault(room_code, {})
    rt["mc_show_correct_answer"] = True
    emit_control_and_mc_state(room_code)

@socketio.on("control_mc_hide_answer")
def control_mc_hide_answer(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    rt = runtime.setdefault(room_code, {})
    rt["mc_show_correct_answer"] = False
    emit_control_and_mc_state(room_code)

@socketio.on("control_show_player_list_on_display")
def control_show_player_list_on_display(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    rt = runtime.setdefault(room_code, {})
    rt["display_show_player_list"] = True
    players = list_players(room_code)
    socketio.emit("show_player_list", {"players": players}, room=f"room:{room_code}:display")
    rt["display_show_qr"] = True
    socketio.emit("show_qr", {}, room=f"room:{room_code}:display")
    emit_control_and_mc_state(room_code)

@socketio.on("control_hide_player_list_on_display")
def control_hide_player_list_on_display(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    rt = runtime.setdefault(room_code, {})
    rt["display_show_player_list"] = False
    rt["display_show_qr"] = False
    socketio.emit("hide_player_list", {}, room=f"room:{room_code}:display")
    socketio.emit("hide_qr", {}, room=f"room:{room_code}:display")
    emit_control_and_mc_state(room_code)


@socketio.on("control_show_not_joined_on_display")
def control_show_not_joined_on_display(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    rt = runtime.setdefault(room_code, {})
    rt["display_show_not_joined"] = True
    not_joined = get_not_joined_list(room_code)
    socketio.emit("show_not_joined_list", {"players": not_joined}, room=f"room:{room_code}:display")
    # Also hide the QR (since it occupies the same space)
    socketio.emit("hide_qr", {}, room=f"room:{room_code}:display")
    emit_control_and_mc_state(room_code)


@socketio.on("control_hide_not_joined_on_display")
def control_hide_not_joined_on_display(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    rt = runtime.setdefault(room_code, {})
    rt["display_show_not_joined"] = False
    socketio.emit("hide_not_joined_list", {}, room=f"room:{room_code}:display")
    emit_control_and_mc_state(room_code)


@socketio.on("control_show_options_on_display")
def control_show_options_on_display(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    if not room_code:
        return
    runtime.setdefault(room_code, {})["display_show_options"] = True
    # cập nhật DB nếu có room
    conn = db()
    cur = conn.cursor()
    cur.execute("UPDATE rooms SET display_show_options=1 WHERE room_code=?", (room_code,))
    conn.commit()
    conn.close()

    socketio.emit("display_state", build_display_state(room_code), room=f"room:{room_code}:display")
    emit_control_and_mc_state(room_code)


@socketio.on("control_hide_options_on_display")
def control_hide_options_on_display(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    if not room_code:
        return
    runtime.setdefault(room_code, {})["display_show_options"] = False
    conn = db()
    cur = conn.cursor()
    cur.execute("UPDATE rooms SET display_show_options=0 WHERE room_code=?", (room_code,))
    conn.commit()
    conn.close()

    socketio.emit("display_state", build_display_state(room_code), room=f"room:{room_code}:display")
    emit_control_and_mc_state(room_code)


# =========================
# SocketIO: player submit
# =========================
@socketio.on("player_submit_answer")
def player_submit_answer(data):
    room_code = str(data.get("room_code", "")).strip().upper()
    player_id = str(data.get("player_id", "")).strip()
    selected_index = int(data.get("selected_index", -1))

    # Nếu player đã bị xóa khỏi DB (hoặc không tồn tại), chặn nộp đáp án
    if not get_player(player_id):
        emit("player_kicked", {"room_code": room_code, "player_id": player_id, "msg": "Bạn đã bị xóa tên, hãy tham gia lại."}, room=request.sid)
        return

    rt = runtime.get(room_code, {})
    # --- 1) Validate logic ---
    if rt.get("status") != "running":
        logger.warning(f"Room {room_code} - Player {player_id} submitted answer but room is NOT running.")
        emit("answer_ack", {"ok": False, "msg": "Phòng chưa bật chế độ thi."}, room=request.sid)
        return
        
    quiz_mode = rt.get("quiz_mode", "host_paced")
    
    if quiz_mode == "self_paced":
        pstate = rt.get("player_states", {}).get(player_id)
        if not pstate:
            emit("answer_ack", {"ok": False, "msg": "Không tìm thấy dữ liệu thi của bạn."})
            return
        if pstate.get("is_finished"):
            emit("answer_ack", {"ok": False, "msg": "Bạn đã hoàn thành bài thi."})
            return
            
        sessions = rt.get("sessions", [])
        q_index = pstate.get("q_index", 0)
        if q_index >= len(sessions):
            return
            
        sess = sessions[q_index]
        session_id = sess["session_id"]
        start_ms = pstate.get("start_ms", now_ms())
        duration = sess["duration"]
        deadline_ms = start_ms + duration * 1000
    else:
        session_id = rt.get("question_session_id")
        q_index = rt.get("question_index", -1)
        start_ms = rt.get("question_start_ms")
        duration = rt.get("duration", 20)
        deadline_ms = int(start_ms) + int(duration)*1000 if start_ms else None

    if not session_id:
        logger.warning(f"Room {room_code} - Player {player_id} submitted answer but NO active question.")
        emit("answer_ack", {"ok": False, "msg": "Không có câu hỏi nào đang mở."}, room=request.sid)
        return
        
    if (quiz_mode == "host_paced" and bool(rt.get("time_up"))) or (deadline_ms and now_ms() > deadline_ms):
        logger.warning(f"Room {room_code} - Player {player_id} submitted answer LATE (Time Up).")
        if quiz_mode == "self_paced":
            pstate["q_index"] += 1
            pstate["start_ms"] = now_ms()
            if pstate["q_index"] >= len(rt.get("sessions", [])):
                pstate["is_finished"] = True
                emit("quiz_ended", {"room_code": room_code}, room=request.sid)
            else:
                emit("question_started", {"q_index": pstate["q_index"]}, room=request.sid)
            return
        else:
            emit("answer_ack", {"ok": False, "msg": "Đã hết thời gian trả lời."}, room=request.sid)
            return

    if has_answer(room_code, session_id, player_id):
        logger.warning(f"Room {room_code} - Player {player_id} attempted to submit MULTIPLE answers.")
        emit("answer_ack", {"ok": False, "msg": "Bạn đã trả lời rồi."}, room=request.sid)
        return

    # Use the q_index properly resolved above depending on quiz_mode
    q = get_question_by_index(room_code, q_index)
    if not q:
        return

    qid = int(q["id"])
    order_str = get_or_create_option_order(room_code, qid, player_id)
    order = list(order_str)  # displayed order of original keys

    if selected_index < 0 or selected_index >= 4:
        emit("answer_ack", {"ok": False, "locked": False, "msg": "Đáp án không hợp lệ"})
        return
    selected_original = order[selected_index]  # A/B/C/D original
    correct = (q["correct"] or "").strip().upper()
    is_correct = 1 if selected_original == correct else 0

    elapsed = max(0, now_ms() - int(start_ms))
    submitted = now_ms()
    
    logger.info(f"Room {room_code} - Player {player_id} submitted Answer: {selected_original} (Correct: {is_correct}) in {elapsed}ms")

    conn = db()
    cur = conn.cursor()
    cur.execute("""
      INSERT INTO answers(room_code, session_id, player_id, selected_original, is_correct, elapsed_ms, submitted_ms, locked)
      VALUES(?,?,?,?,?,?,?,1)
    """, (room_code, session_id, player_id, selected_original, is_correct, int(elapsed), int(submitted)))
    conn.commit()
    conn.close()

    if quiz_mode == "self_paced":
        pstate["q_index"] += 1
        pstate["start_ms"] = now_ms()
        if pstate["q_index"] >= len(rt.get("sessions", [])):
            pstate["is_finished"] = True
            emit("quiz_ended", {"room_code": room_code}, room=request.sid)
        else:
            emit("question_started", {"q_index": pstate["q_index"]}, room=request.sid)
            
        emit("answer_ack", {"ok": True, "locked": True, "is_correct": bool(is_correct)}, room=request.sid)
        emit_control_and_mc_state(room_code)
        socketio.emit("display_state", build_display_state(room_code), room=f"room:{room_code}:display")
    else:
        emit("answer_ack", {"ok": True, "locked": True, "is_correct": bool(is_correct)})
        emit_control_and_mc_state(room_code)


# =========================
# Flow control: next question
# =========================
def start_self_paced_quiz(room_code: str):
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT id, duration_sec FROM questions WHERE room_code=? ORDER BY CAST(stt AS INTEGER)", (room_code,))
    qs = cur.fetchall()
    
    sessions = []
    import uuid
    import time
    for i, q in enumerate(qs):
        sid = str(uuid.uuid4())
        dur = int(q["duration_sec"])
        cur.execute("""
            INSERT INTO question_sessions(id, room_code, q_index, question_id, start_ms, duration)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (sid, room_code, i, q["id"], int(time.time() * 1000), dur))
        sessions.append({
            "session_id": sid,
            "duration": dur,
            "question_id": q["id"]
        })
    conn.commit()
    
    rt = runtime.get(room_code, {})
    rt["sessions"] = sessions
    rt["player_states"] = {}
    
    cur.execute("SELECT id FROM players WHERE room_code=?", (room_code,))
    players = cur.fetchall()
    conn.close()
    
    now_ms = int(time.time() * 1000)
    for p in players:
        rt["player_states"][p["id"]] = {
            "q_index": 0,
            "start_ms": now_ms,
            "is_finished": False
        }
        
    socketio.emit("question_started", {"q_index": 0}, room=f"room:{room_code}:players")
    socketio.emit("question_started", {"q_index": 0}, room=f"room:{room_code}:display")
    socketio.emit("question_started", {"q_index": 0}, room=f"room:{room_code}:mc")
    
    emit_control_and_mc_state(room_code)
    socketio.emit("display_state", build_display_state(room_code), room=f"room:{room_code}:display")
    self_paced_timer_daemon(room_code)

def start_next_question(room_code: str):
    rt = runtime.get(room_code)
    if not rt:
        set_runtime_from_db(room_code)
        rt = runtime.get(room_code)

    if rt.get("status") != "running":
        return

    prev_sid = rt.get("question_session_id")
    if prev_sid:
        end_question_session(prev_sid)
        stats = compute_question_stats(room_code, prev_sid)
        top10 = compute_scoreboard(room_code)[:10]
        socketio.emit("question_ended", {"stats": stats, "top10": top10}, room=f"room:{room_code}:control")
        socketio.emit("question_ended", {}, room=f"room:{room_code}:display")
        socketio.emit("question_ended", {}, room=f"room:{room_code}:players")
        socketio.emit("question_ended", {}, room=f"room:{room_code}:mc")

    next_index = int(rt.get("question_index", -1)) + 1
    q_total = total_questions(room_code)
    if next_index >= q_total:
        logger.info(f"Room {room_code} - Auto-ended quiz (reached last question).")
        conn = db()
        cur = conn.cursor()
        cur.execute("UPDATE rooms SET status='ended' WHERE room_code=?", (room_code,))
        conn.commit()
        conn.close()

        # Flush deferred connection logs
        flush_connection_logs(room_code)

        rt["status"] = "ended"
        socketio.emit("quiz_ended", {"room_code": room_code}, room=f"room:{room_code}:display")
        socketio.emit("quiz_ended", {"room_code": room_code}, room=f"room:{room_code}:players")
        socketio.emit("quiz_ended", {"room_code": room_code}, room=f"room:{room_code}:mc")
        for p in list_players(room_code):
            pid = p.get("id")
            if pid:
                socketio.emit("player_final_stats", compute_player_final_stats(room_code, pid), room=f"player:{pid}")
        emit_control_and_mc_state(room_code)
        return

    q = get_question_by_index(room_code, next_index)
    if not q:
        return

    room = get_room(room_code)
    fallback = int(room["duration"]) if room else 20
    qdur = int(q["duration_sec"]) if ("duration_sec" in q.keys() and q["duration_sec"] is not None) else 0
    duration = qdur if qdur > 0 else fallback
    sid = create_question_session(room_code, next_index, int(q["id"]), duration)
    start_ms = now_ms()

    rt["question_index"] = next_index
    rt["question_session_id"] = sid
    rt["question_start_ms"] = start_ms
    rt["duration"] = duration
    rt["time_up"] = False
    rt["reveal_on"] = False
    rt["display_show_winner"] = False

    logger.info(f"Room {room_code} - Started question session: {sid} (Q_index: {next_index})")

    socketio.emit("hide_answer", {}, room=f"room:{room_code}:display")
    socketio.emit("hide_stats", {}, room=f"room:{room_code}:display")

    socketio.emit("question_started", build_display_state(room_code), room=f"room:{room_code}:display")
    socketio.emit(
        "question_started",
        {"status": "running", "q_index": next_index, "start_ms": start_ms, "duration": duration},
        room=f"room:{room_code}:players"
    )
    socketio.emit(
        "question_started",
        {"status": "running", "q_index": next_index, "start_ms": start_ms, "duration": duration},
        room=f"room:{room_code}:mc"
    )
    emit_control_and_mc_state(room_code)

    schedule_time_up(room_code, sid, duration)


# =========================
# Control: delete player (before start only)
# =========================
@socketio.on("control_delete_player")
def control_delete_player(data):
    if not session.get("is_admin"):
        return
    room_code = str(data.get("room_code", "")).strip().upper()
    player_id = str(data.get("player_id", "")).strip()
    if not room_code or not player_id:
        return

    room = get_room(room_code)
    if not room:
        emit("toast", {"type": "error", "msg": "ROOM không tồn tại."}, room=request.sid)
        return


    # kiểm tra tồn tại
    player = get_player(player_id)
    if not player:
        emit("toast", {"type": "warn", "msg": "Người chơi không tồn tại (có thể đã bị xóa)."}, room=request.sid)
        return

    conn = db()
    cur = conn.cursor()
    
    # Lấy thông tin player để đưa vào deleted_players
    deleted_at = datetime.now(timezone.utc).isoformat()
    cur.execute("""
      INSERT INTO deleted_players (id, room_code, name, position, unit, joined_at, employee_code, deleted_at)
      VALUES (?, ?, ?, ?, ?, ?, ?, ?)
      ON CONFLICT(id) DO UPDATE SET deleted_at=excluded.deleted_at
    """, (player["id"], player["room_code"], player["name"], player["position"], player["unit"], player["joined_at"], player["employee_code"], deleted_at))

    # xóa dữ liệu liên quan (nếu có)
    cur.execute("DELETE FROM answers WHERE room_code=? AND player_id=?", (room_code, player_id))
    cur.execute("DELETE FROM option_maps WHERE room_code=? AND player_id=?", (room_code, player_id))
    cur.execute("DELETE FROM players WHERE room_code=? AND id=?", (room_code, player_id))
    conn.commit()
    conn.close()

    logger.info(f"Room {room_code} - HOST deleted player {player_id}")

    # đá người chơi khỏi màn Player (client sẽ tự redirect về join)
    socketio.emit(
        "player_kicked",
        {"room_code": room_code, "player_id": player_id, "msg": "Bạn đã bị xóa tên, hãy tham gia lại."},
        room=f"player:{player_id}"
    )

    players = list_players(room_code)
    socketio.emit("player_list_updated", {"players": players}, room=f"room:{room_code}:control")
    socketio.emit("player_list_updated", {"players": players}, room=f"room:{room_code}:display")
    socketio.emit("player_list_updated", {"players": players}, room=f"room:{room_code}:mc")
    emit_control_and_mc_state(room_code)

# =========================
# Main
# =========================
init_db()  # Ensure database is initialized even when imported by gunicorn

if __name__ == "__main__":
    socketio.run(app, host="0.0.0.0", port=5000, debug=True, allow_unsafe_werkzeug=True)
